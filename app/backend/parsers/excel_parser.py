"""Excel parsing service."""
import logging
from typing import Dict, Any, List
import pandas as pd
from openpyxl import load_workbook
from google.cloud import storage

from config import settings

logger = logging.getLogger(__name__)


class ExcelParser:
    """Parse Excel files and extract financial data."""
    
    def __init__(self):
        self.storage_client = storage.Client(project=settings.project_id)
    
    def parse_excel(self, gcs_path: str) -> Dict[str, Any]:
        """
        Parse an Excel file.
        
        Args:
            gcs_path: GCS path to the Excel file
            
        Returns:
            Parsed Excel data with sheets and tables
        """
        logger.info(f"Parsing Excel file: {gcs_path}")
        
        try:
            # Download file from GCS
            bucket = self.storage_client.bucket(settings.uploads_bucket)
            blob = bucket.blob(gcs_path)
            
            # Download to temp location
            temp_file = f"/tmp/{blob.name.split('/')[-1]}"
            blob.download_to_filename(temp_file)
            
            # Load workbook
            wb = load_workbook(temp_file, data_only=True)
            
            # Extract data from each sheet
            sheets = {}
            for sheet_name in wb.sheetnames:
                sheet_data = self._parse_sheet(temp_file, sheet_name)
                if sheet_data:
                    sheets[sheet_name] = sheet_data
            
            return {
                "sheets": list(sheets.keys()),
                "data": sheets,
                "total_sheets": len(wb.sheetnames)
            }
            
        except Exception as e:
            logger.error(f"Error parsing Excel file {gcs_path}: {str(e)}")
            raise
    
    def _parse_sheet(self, file_path: str, sheet_name: str) -> Dict[str, Any]:
        """Parse a single Excel sheet."""
        try:
            # Read with pandas for easier data manipulation
            df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
            
            # Find header row (first non-empty row with mostly text)
            header_row = self._find_header_row(df)
            
            if header_row is not None:
                # Re-read with proper header
                df = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row)
                
                # Clean column names
                df.columns = [str(col).strip() for col in df.columns]
                
                # Drop empty rows and columns
                df = df.dropna(how='all').dropna(axis=1, how='all')
                
                # Detect numeric columns (likely amounts)
                numeric_cols = df.select_dtypes(include=['float64', 'int64']).columns.tolist()
                
                return {
                    "header_row": header_row,
                    "columns": df.columns.tolist(),
                    "numeric_columns": numeric_cols,
                    "rows": len(df),
                    "data": df.to_dict(orient='records')
                }
            
            return None
            
        except Exception as e:
            logger.warning(f"Error parsing sheet {sheet_name}: {str(e)}")
            return None
    
    def _find_header_row(self, df: pd.DataFrame, max_search_rows: int = 20) -> int:
        """
        Identify the header row by finding the first row with mostly text values.
        
        Args:
            df: DataFrame to search
            max_search_rows: Maximum number of rows to search
            
        Returns:
            Row index of header, or None if not found
        """
        for i in range(min(max_search_rows, len(df))):
            row = df.iloc[i]
            
            # Count non-empty cells
            non_empty = row.notna().sum()
            
            if non_empty >= 2:  # At least 2 column headers
                # Check if mostly text (not numbers)
                text_count = sum(isinstance(val, str) for val in row if pd.notna(val))
                
                if text_count >= non_empty * 0.5:  # At least 50% text
                    return i
        
        return None
    
    def detect_financial_statement_type(self, sheet_data: Dict[str, Any]) -> str:
        """
        Detect the type of financial statement from sheet data.
        
        Args:
            sheet_data: Parsed sheet data
            
        Returns:
            Statement type: income_statement, balance_sheet, cash_flow, or other
        """
        if not sheet_data or 'data' not in sheet_data:
            return "other"
        
        # Get first column (usually line item names)
        df = pd.DataFrame(sheet_data['data'])
        if df.empty:
            return "other"
        
        first_col = df.iloc[:, 0].astype(str).str.lower()
        all_text = ' '.join(first_col)
        
        # Income statement indicators
        if any(keyword in all_text for keyword in ['revenue', 'sales', 'cogs', 'gross profit', 'net income', 'operating income']):
            return "income_statement"
        
        # Balance sheet indicators
        if any(keyword in all_text for keyword in ['assets', 'liabilities', 'equity', 'retained earnings', 'accounts receivable']):
            return "balance_sheet"
        
        # Cash flow indicators
        if any(keyword in all_text for keyword in ['cash flow', 'operating activities', 'investing activities', 'financing activities']):
            return "cash_flow"
        
        return "other"

